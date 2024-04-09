import asyncio

from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook
from service.main_test import sync_main_test
from tg_bot.utils.utils import api_send_memory, api_send_result
from service import flags


# функция очистки файла result_work.xlsx (запуск еженедельно)
def _cleaner_file():
    wb = load_workbook("result_work.xlsx")
    ws = wb['Лист1']
    max_rows = ws.max_row
    for i in range(max_rows, 2, -1):
        if i >= 3:
            ws.delete_rows(i)
    wb.save("result_work.xlsx")


# функция планировщик
async def plan_work():
    scheduler_1 = BackgroundScheduler()
    scheduler_1.add_job(api_send_memory, 'cron', hour=12, misfire_grace_time=60 * 5, max_instances=1)
    scheduler_1.add_job(api_send_result, 'cron', day_of_week='mon', hour=10, misfire_grace_time=60 * 5,
                        max_instances=1)
    scheduler_1.add_job(_cleaner_file, 'cron', day_of_week='mon', hour=10, minute=5, misfire_grace_time=60 * 5,
                        max_instances=1)
    scheduler_1.start()
    scheduler_2 = BackgroundScheduler()
    job_1 = scheduler_2.add_job(
        sync_main_test, 'cron', hour=7, minute=30, args=[0], misfire_grace_time=60 * 5, max_instances=1
    )
    job_2 = scheduler_2.add_job(
        sync_main_test, 'cron', hour=16, minute=30, args=[1], misfire_grace_time=60 * 5, max_instances=1
    )
    scheduler_2.start()

    j = 1
    while flags.flag_work_script:
        if flags.flag_scheduled_check and not j:
            scheduler_2.resume_job(job_1.id)
            scheduler_2.resume_job(job_2.id)
            j += 1
        elif not flags.flag_scheduled_check and j:
            scheduler_2.pause_job(job_1.id)
            scheduler_2.pause_job(job_2.id)
            j -= 1
        await asyncio.sleep(5)
    scheduler_1.shutdown()
    scheduler_2.shutdown()

    # Инициализируем исключение для выключения программы,
    # если флаг flag_work_script пользователем переведён в False
    raise Exception('The program was stopped by the user')
