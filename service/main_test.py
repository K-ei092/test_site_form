from checking_mail import checking_mail
import datetime
import asyncio
from selenium_test.selenium_test_site import SiteTest
from openpyxl import load_workbook
from tg_bot.utils.utils import api_send_message
from service import flags


async def _fil_table(name_test: str, result_post: tuple) -> None:
    wb = load_workbook("result_work.xlsx")
    ws = wb.active
    num = int(name_test[12:])
    ws.cell(row=num + 2, column=1, value=name_test)
    result_post = list(result_post)
    if (result_post[0] == 'Сайт не дал ответа') or (200 <= int(result_post[0]) <= 399):
        result_post[2] = (datetime.datetime.strptime(result_post[2], '%m/%d/%y %H:%M:%S')
                          - datetime.timedelta(seconds=5)).strftime('%m/%d/%y %H:%M:%S')
        ws.cell(row=num + 2, column=2,
                value=f'{result_post[1]} (ответ сайта - {str(result_post[0])})')  # вносим название формы в базу
        ws.cell(row=num + 2, column=3, value=result_post[2])

        result_mail = checking_mail.get_mail(name_test)  # проверка почты и возврат времени поступления письма

        if isinstance(result_mail, str):
            time1 = datetime.timedelta(hours=int(result_post[2][-8:-6]), minutes=int(result_post[2][-5:-3]),
                                       seconds=int(result_post[2][-2:]))
            time2 = datetime.timedelta(hours=int(result_mail[-8:-6]), minutes=int(result_mail[-5:-3]),
                                       seconds=int(result_mail[-2:]))
            result_time = str(time2 - time1)  # получаем разницу времени отправки на сайте и поступления на почту
            ws.cell(row=num + 2, column=4, value=result_mail)
            ws.cell(row=num + 2, column=5, value=result_time)
            ws.cell(row=num + 2, column=6, value=result_post[3])
            ws.cell(row=num + 2, column=7, value='-------')
        else:
            api_send_message(text=f'Форма "{result_post[1]}" - проверка почты: {result_mail[1]}')
            ws.cell(row=num + 2, column=4, value=f'проверка почты: {result_mail[1]}')
            ws.cell(row=num + 2, column=5, value='-------')
            ws.cell(row=num + 2, column=6, value=result_post[3])
            ws.cell(row=num + 2, column=7, value=datetime.datetime.now())

    else:
        api_send_message(text=f'Форма "{result_post[1]}" - ошибка заполнения: {result_post[2]}')
        ws.cell(row=num + 2, column=2, value=result_post[1])
        ws.cell(row=num + 2, column=3, value=result_post[2])
        ws.cell(row=num + 2, column=4, value='-------')
        ws.cell(row=num + 2, column=5, value='-------')
        ws.cell(row=num + 2, column=6, value=result_post[3])
        ws.cell(row=num + 2, column=7, value=datetime.datetime.now())

    wb.save("result_work.xlsx")
    await asyncio.sleep(5)


async def main_test(ua):
    if not flags.flag_process_started:
        flags.flag_process_started = True

        wb = load_workbook("result_work.xlsx")
        ws = wb.active  # активируем лист таблицы xlsx
        num = ws['A'][-1].value  # формируем имя очередного теста из предыдущего, записанного в базе
        num = int(num[12:]) + 1

        name_test_1 = f'Number_TEST_{str(num)}'
        st = SiteTest(n=ua)
        form_mm_novuy = st.check_mm_novuy_2(name_test_1)
        await asyncio.sleep(3)

        name_test_2 = f'Number_TEST_{str(num + 1)}'
        st = SiteTest(n=ua)
        form_zakazat_zvonok = st.check_zakazat_zvonok_3(name_test_2)
        await asyncio.sleep(3)

        name_test_3 = f'Number_TEST_{str(num + 2)}'
        st = SiteTest(n=ua)
        form_consultation = st.check_consultation_4(name_test_3)
        await asyncio.sleep(3)

        await asyncio.sleep(20 * 60)

        await _fil_table(name_test_1, form_mm_novuy)
        await _fil_table(name_test_2, form_zakazat_zvonok)
        await _fil_table(name_test_3, form_consultation)

        flags.flag_process_started = False

        if ua == 0:
            api_send_message(text='тестирование завершено (моб. версия)')
        else:
            api_send_message(text='тестирование завершено (ПК версия)')


def sync_main_test(ua):
    asyncio.run(main_test(ua))


if __name__ == '__main__':
    main_test()
